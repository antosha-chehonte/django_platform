import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def export_to_csv(data, filename, headers, row_generator):
    """
    Экспорт данных в CSV
    
    Args:
        data: данные для экспорта
        filename: имя файла
        headers: список заголовков колонок
        row_generator: функция-генератор строк (принимает элемент данных, возвращает список значений)
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for item in data:
        row = row_generator(item)
        writer.writerow(row)
    
    return response


def export_to_excel(data, filename, headers, row_generator, sheet_name='Отчет'):
    """
    Экспорт данных в Excel
    
    Args:
        data: данные для экспорта
        filename: имя файла
        headers: список заголовков колонок
        row_generator: функция-генератор строк
        sheet_name: название листа
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for row_num, item in enumerate(data, 2):
        row_data = row_generator(item)
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top')
    
    # Автоматическая ширина колонок
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    wb.save(response)
    return response

